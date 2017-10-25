# -*- coding: utf-8 -*-
"""
Created on Fri Sep 29 14:53:36 2017
This file combine all other Python files with some improvments to avoid useless processes
This is an optimization of the Home Depot .com Delivery network, with objective function being the total cost (line haul + last Mile) 
@author: Cyprien Bastide, Steven (Gao) Ming, Edson David Silva Moreno
"""



#Import the differents Modules that are going to be used in this file
import sys
#   This Module is used to open excel files
import openpyxl as xl
# Neig_states return the neighboring state of the input state
# It's an easier way to call cell in an excel file
# instance return the number of lines in an excel spreadsheet
# Compute distance2 compute the distances from 2 zip code and the lat long database
# Correct zip add 0 in front of postal codes that are not 5 digits long
# Get last mile pricing returns a dictionnary containing the info to compute the last mile cost
from Procedures import neig_states, cell, instance, compute_distance2, correct_zip, get_lm_pricing, averageOrig
#tqdm is used to create progress bar, however, if computation is too fast it might cause bugs
from tqdm import tqdm
# Pulp is the optimization engine
import pulp
#time allow to compute time elapsed for some taks
import time
import pandas as pd
number_days = 30*6
weight_treshold_ltl = 200
nb_trucks = round(number_days*5/7)
weight_per_volume = 100
coefficient= {'intercept':-3.683,'weight':0.1498,'dist':0.0537,'weight_dist':0.0001,'CA':0,"GA":-8.4855,"MD":-7.5867,"OH":3.4399}

"""
###############################################################
###############################################################

This part upload the information from the excel spreadsheet

###############################################################
###############################################################
"""
print('Import Workbook') 
# Open Worksheet that contains list of DA, List of Zip code with Volume, Pricing spreadsheet,...
wb = xl.load_workbook('C:\HomeDepot_Excel_Files\Standard_File.xlsx')
# Open All Different Spreadsheet
w_neig = wb['List_of_Neighboring_States']
w_da = wb['DA_List']
w_zip = wb['Zip_Allocation_and_Pricing']
w_lm_pricing = wb["LM_Pricing"]
w_range = wb["Zip_Range"]
w_dfc = wb["DFC list"]
# Import Database of Zipcode Latitude and Longitude
print('Open Lat Long Database')
wdata = xl.load_workbook('C:\HomeDepot_Excel_Files\Zip_latlong.xlsx')
wslatlong = wdata['Zip']

#Importing Excel sheet as Panda Data Frames to create Dictionary with every destination state
#as a Key and each Key has a nested Dictionary with the weight (percentage) of invoices 
#coming from every origin for LTL pricing.
print('Import Database LTL')
wbLtl = pd.ExcelFile('C:\HomeDepot_Excel_Files\ltl_price.xlsx')
ltl_price = wbLtl.parse('ltl_price', converters={'dest_zip': str,'orig_zip': str})


"""
###############################################################
###############################################################

This part create different dictionnaries that are going to be used to solve the optimization problem
Multiple dictionnaries are going to be created to represent DAs and Zipcode because we will need different 
keys to call them

###############################################################
###############################################################
"""

print ("Create all Dictionnaries")
# Get number of DA and their Zip Codes
n_da= instance(w_da)
n_zip = instance(w_zip)
linelatlong = instance(wslatlong)
# Create dictionnaries for DA and Zip(with volume in tuple) Grouped by State, 
# This is useful because Arcs are going to be created based on neighgboring states
# Since distances don't depend on Carrier, if multiple DA are in the same zipcode only one will be counted

# Dictionnary for DA  {State : [Da_ZipCode]} 
State_Da_dict = {}
for r in range(n_da):
    state = cell(w_da,r+2,3) 
    da_zip = correct_zip(str(cell(w_da,r+2,2))) 
#        Remove duplicate: DA in same zipcode but different carriers
    State_Da_dict[state] = list(set().union(State_Da_dict.setdefault(state,[]),[da_zip]))


# Dictionnary for Zip {State : [( zipcode, volume)]}
State_Zip_dict = {}
for r in range(n_zip):
    state = cell(w_zip,r+2,2)
    zipcode = correct_zip(str(cell(w_zip,r+2,1))) 
    volume = cell(w_zip,r+2,7)
    
    State_Zip_dict.setdefault(state, []).append((zipcode, volume))

#Create Dictionnary for Da { Zip:(Zip, State, [Carrier])} will be useful o compute distances

DA_ZipCode_Dict = {}
for r in range(n_da):
    zipcode = correct_zip(str(cell(w_da,r+2,2)))
    state = cell(w_da, r+2, 3)
    carrier = cell(w_da,2+r, 4)
    
    DA_ZipCode_Dict.setdefault(zipcode,{'Zip':zipcode, 'State':state, 'Carrier':[carrier]}) 

    DA_ZipCode_Dict[zipcode]['Carrier'] = list(set().union(DA_ZipCode_Dict[zipcode]['Carrier'],[carrier]))



# Other dictionnary for Da, { Zip + Carrier : (Zip, State, Carrier)}

DAC_ZipCode_Dict = {}
for r in range(n_da):
    zipcode = correct_zip(str(cell(w_da,r+2,2)))
    carrier = cell(w_da, r+2, 4)
    state = cell(w_da, r+2, 3)
    DAC_ZipCode_Dict[zipcode+' '+ carrier] = {'Zip':zipcode, 'State':state, 'Carrier':carrier}


#Create Dictionnary for Zipcode (Zip, Volume ,State)
ZipCode_Dict={}
for r in range(n_zip):
    zipcode = correct_zip(str(cell(w_zip,2+r, 1)))
    volume = cell(w_zip, r+2, 7)
    state = cell(w_zip,r+2, 2)
    ZipCode_Dict[zipcode]={'Zip':zipcode,'Volume':volume, 'State':state}         

# Create LM Pricing Dictionnary
Pricing=get_lm_pricing(w_lm_pricing)

# Get arc max range Dictionnary {State : Max range}
Range = { cell(w_range,r+2,2) : cell(w_range,r+2,3) for r in range(instance(w_range))}


# Create dictionnary for the database lat long{Zip : (lat,long)}
Zip_lat_long = {}
for r in range(linelatlong):
    zipcode = correct_zip(str(cell(wslatlong,r+2,1)))
    lat = cell(wslatlong,r+2,2)
    long = cell(wslatlong,r+2,3)
    Zip_lat_long[zipcode] = (lat,long)

#Create dictionary with State Destination as a Key and nested dictionary with weight by origin 
percDestin = averageOrig(ltl_price)

# Create Dictionnary of DFC {State : {Name, Zip, State}}
DFC_Dict={}
nbdfc = instance(w_dfc)
for r in range(nbdfc):
    name = cell(w_dfc,r+2,1)
    zipcode = correct_zip(str(cell(w_dfc,r+2,2)))
    state = cell(w_dfc,r+2,3)
    DFC_Dict[state]={'State':state,'Name':name,'Zipcode':zipcode}
"""
###############################################################
###############################################################

This part compute the distance between Das and DFC and the pricing based on parameters and distance

\\ Problem for new DAs in States without Das currently (Key Error)

###############################################################
############################################################### 
"""
# Relation Da_Dfc is in format {da : {state_dfc : {distance, percentage from state}, global : {slope,cost_opening}}
a = 0
Da_Dfc = {}

for da in DA_ZipCode_Dict.keys():
    
    da_state = DA_ZipCode_Dict[da]["State"]
    try:
        for dfc_state in percDestin[da_state].keys():
            
            dfc_zip = DFC_Dict[dfc_state]['Zipcode']
            percentage = percDestin[da_state][dfc_state]
            
            distance, Zip_lat_long, b = compute_distance2(da,dfc_zip,Zip_lat_long)
            
            slope = coefficient["weight"]+coefficient["weight_dist"]*distance
            
            intercept = coefficient['intercept']+coefficient[dfc_state]+coefficient['dist']*distance
            
            cost_opening = intercept + weight_treshold_ltl * slope
            
            Da_Dfc.setdefault(da,{}).setdefault(dfc_state, {"distance":distance, "percentage" : percentage,"slope": slope,"cost_opening": cost_opening})
            if b == 1 :
                a += 1
        slope = sum(Da_Dfc[da][dfc_state]["slope"]*Da_Dfc[da][dfc_state]["percentage"] for dfc_state in Da_Dfc[da].keys())
        cost_opening = sum(Da_Dfc[da][dfc_state]["cost_opening"]*Da_Dfc[da][dfc_state]["percentage"] for dfc_state in Da_Dfc[da].keys())
        Da_Dfc[da]['Global']={'slope' : slope, 'cost_opening' : cost_opening}
    except KeyError:
#        Just assume we take the previous cost
        Da_Dfc.setdefault(da,{}).setdefault('Global',{'slope' : slope, 'cost_opening' : cost_opening, 'Warning':"This Da doesn't have real slope or cost of opening"}  )  

"""
###############################################################
###############################################################

This part create the combination DA Zipcode based on neighboring states and then compute the distances 

###############################################################
###############################################################
"""

# Create couple that need to have distance assigned to it based Neighboring states, 
# Go through every state, look at the Da inside, and assign them to all zip code in neighboring states

combination = []
# Iterate through the states
print('Creating combination Da Zipcode')
for state in tqdm(State_Da_dict.keys()):
#    Create list of Da in the state
    da_list = State_Da_dict[state]

#   Create list of Zipcode that are in neigh state
    zip_list = []
    neighboring_states= neig_states(state,w_neig)
    for n_state in neighboring_states:
        zip_list += State_Zip_dict[n_state]
# Create combination if Volume is not 0, have list of all combination [[da,zip]]
    for da in da_list:
        for pc in zip_list:
            zipcode = pc[0]
            volume = pc[1]
            if volume != 0:
                combination.append([da,zipcode])
                
# Compute distances for each combination
nb_distances = len(combination)

print("Compute distances")
for i in tqdm(range(nb_distances)):
    da = combination[i][0]
    zipcode = combination[i][1] 
    distance, Zip_lat_long, b = compute_distance2(da, zipcode, Zip_lat_long)
    combination[i].append(distance)
    if b == 1 :
        a += 1

# Update if new zipcodes have been added
if a != 0:
    print("Update Database")
    ZipList = Zip_lat_long.keys()
    c = 0
    for r in ZipList:
        wslatlong.cell(row = c+2,column = 1).value = r
        wslatlong.cell(row = c+2,column = 2).value = Zip_lat_long[r][0]
        wslatlong.cell(row = c+2,column = 3).value = Zip_lat_long[r][1]
        c+=1
    wdata.save('C:\HomeDepot_Excel_Files\Zip_latlong.xlsx')
    print('Database updated')
    
    
"""
###############################################################
###############################################################

This part is the first Optimization model (only last mile) that will remove DA
that are useless to achieve faster computation time for the last mile and line haul optimization

###############################################################
###############################################################
"""

#Create arcs out of combination if distance is less than threshold 
#   { Zip : {DA+Carrier :{distance, lm_cost (come in next step), var(come in two steps)}}
print("Build Arcs")

Arcs={}
for i in tqdm(range(nb_distances)):
    da = combination[i][0]
    zipcode = combination[i][1] 
    distance = combination[i][2]
#        Create an arc only if distance between DA and Zip is less than the Zip's state threshold
    if distance< Range[ZipCode_Dict[zipcode]['State']]:
        try:
            for carrier in DA_ZipCode_Dict[da]['Carrier']:
                Arcs[zipcode][da +" "+carrier ]={'distance' : distance}
        except KeyError:
            Arcs[zipcode]= {}
            for carrier in DA_ZipCode_Dict[da]['Carrier']:
                 Arcs[zipcode][da + " "+carrier]={'distance' : distance}

# Compute Costs for the arcs
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        
        distance = Arcs[pc][da]['distance']        
        da_state = DAC_ZipCode_Dict[da]['State']
        da_carrier = DAC_ZipCode_Dict[da]['Carrier']
        
        try:
            flat = Pricing[da_state][da_carrier]['Flat']
            breakpoint = Pricing[da_state][da_carrier]['Break']
            extra = Pricing[da_state][da_carrier]['Extra']
        except KeyError:
            sys.exit(("LM_Cost spreadsheet does not contain pricing info for couple state-carrier %s, %s" %(da_state,da_carrier)))
            
#        Check if distance Da_Zip is within flat distance
        if distance <  breakpoint:
            Arcs[pc][da]['lm_cost']=flat
        else :
            Arcs[pc][da]['lm_cost']=flat+ (distance - breakpoint) * extra
            
# Create Model
prob = pulp.LpProblem("Minimize Distance",pulp.LpMinimize)

# Design arcs
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        var = pulp.LpVariable("Arc_%s_%s)" % (pc,da),0,1,pulp.LpContinuous)
        Arcs[pc][da]['variable']=var

 

# Create Objective function : minimize distance
print("Create objective and Constraint")

#           We add a fraction of distance to the lm cost so we can avoid equality in price
prob += pulp.lpSum([(Arcs[pc][da]['lm_cost']+0.001*Arcs[pc][da]['distance'])*Arcs[pc][da]['variable'] for pc in Arcs.keys() for da in Arcs[pc].keys()])

# Create Constraint : every Zip is allocated
print("Create contraint 'every zipcode is assigned to a DA'")
for pc in tqdm(Arcs.keys()):          
    prob += pulp.lpSum([Arcs[pc][da]['variable'] for da in Arcs[pc].keys()]) == 1

# The problem is solved using PuLP's choice of Solver
print("Solve Problem")
start_time = time.clock()
prob.solve()
end_time = time.clock()
print(end_time-start_time)

# The status of the solution is printed to the screen
print("Status:", pulp.LpStatus[prob.status])


# The optimised objective function value is printed to the screen    
print ("total cost", pulp.value(prob.objective))

##Create workbook for results
#w_result = xl.Workbook()
#wresult = w_result.create_sheet('Optimization Results')
## export results on excel

#print("Exporting Results")
#
#wresult.cell(row=1,column=1).value= "ZipCode"
#wresult.cell(row=1,column=2).value= "Carrier"
#wresult.cell(row=1,column=3).value= "DaZipCode"
#wresult.cell(row=1,column=4).value= 'DA and Carrier'
#wresult.cell(row=1,column=5).value= 'Volume'
#wresult.cell(row=1,column=6).value= 'Unit Cost'
## Print Results on excel
#r=2
#for pc in Arcs.keys():
#    for da in Arcs[pc].keys():
#        if Arcs[pc][da]['variable'].varValue !=0:
#            wresult.cell(row=r,column=1).value= pc
#            wresult.cell(row=r,column=2).value= da[6:]
#            wresult.cell(row=r,column=3).value= da[:5]
#            wresult.cell(row=r,column=4).value= da
#            wresult.cell(row=r,column=5).value= ZipCode_Dict[pc]['Volume']            
#            wresult.cell(row=r,column=6).value= Arcs[pc][da]['lm_cost']
#            r+=1
#
#
#print("Save File")
#
#w_result.save("C:\HomeDepot_Excel_Files\Optimized.xlsx")

# Return List of useful DA
Useful_Da = []
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        if Arcs[pc][da]['variable'].varValue != 0:
            Useful_Da = list(set().union([da],Useful_Da))


  
"""
###############################################################
###############################################################

This part is the second optimization model (includes last mile and line haul)
and uses only DAs that are useful (based on previous optimization)

\For now arcs are recreated while we could use previous dictionnary
\Model only has one treshold
###############################################################
###############################################################
"""
# Remove useless arcs
Arcs={}
for i in tqdm(range(nb_distances)):
    da = combination[i][0]
    zipcode = combination[i][1] 
    distance = combination[i][2]
#        Create an arc only if distance between DA and Zip is less than the Zip's state threshold
    if distance< Range[ZipCode_Dict[zipcode]['State']]:
        try:
            for carrier in DA_ZipCode_Dict[da]['Carrier']:
                if da +" "+carrier in Useful_Da:
                    Arcs[zipcode][da +" "+carrier ]={'distance' : distance}
        except KeyError:
            Arcs[zipcode]= {}
            for carrier in DA_ZipCode_Dict[da]['Carrier']:
                if da +" "+carrier in Useful_Da:
                    Arcs[zipcode][da + " "+carrier]={'distance' : distance}

# Compute Costs for the arcs
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        
        distance = Arcs[pc][da]['distance']        
        da_state = DAC_ZipCode_Dict[da]['State']
        da_carrier = DAC_ZipCode_Dict[da]['Carrier']
        
        try:
            flat = Pricing[da_state][da_carrier]['Flat']
            breakpoint = Pricing[da_state][da_carrier]['Break']
            extra = Pricing[da_state][da_carrier]['Extra']
        except KeyError:
            sys.exit(("LM_Cost spreadsheet does not contain pricing info for couple state-carrier %s, %s" %(da_state,da_carrier)))
            
#        Check if distance Da_Zip is within flat distance
        if distance <  breakpoint:
            Arcs[pc][da]['lm_cost']=flat
        else :
            Arcs[pc][da]['lm_cost']=flat+ (distance - breakpoint) * extra
            

#   { Zip : {DA+Carrier :{distance, lm_cost (come in next step), var(come in two steps)}}
print("Build Arcs")

# Create Model
prob = pulp.LpProblem("Minimize HDU Cost",pulp.LpMinimize)

# Design arcs
for pc in Arcs.keys():
    for da in Arcs[pc].keys():

        var = pulp.LpVariable("Arc_%s_%s)" % (pc,da),0,1,pulp.LpContinuous)
        Arcs[pc][da]['variable']=var
            
# Create variable for Das (OPEN AND VOLUME OVER 200)
for da in Useful_Da:

    ovar = pulp.LpVariable("Da_%s" % (str(da)),0,1,pulp.LpBinary)
    wvar = pulp.LpVariable("Da_%s_above_200LBS" % (str(da)))
    DAC_ZipCode_Dict[da]['opening_variable']=ovar
    DAC_ZipCode_Dict[da]['Weight_variable']=wvar


# Create Objective function : minimize distance
print("Create objective and Constraint")
def lmcost(pc,da):
    return Arcs[pc][da]['lm_cost']*ZipCode_Dict[pc]['Volume']*Arcs[pc][da]['variable']
def lhcost(da):
    zip_da = da[:5]
    return Da_Dfc[zip_da]['Global']['cost_opening']*DAC_ZipCode_Dict[da]['opening_variable'] + Da_Dfc[zip_da]['Global']['slope'] * DAC_ZipCode_Dict[da]["Weight_variable"]

prob += pulp.lpSum([lmcost(pc,da) for pc in Arcs.keys() for da in Arcs[pc].keys()]) + nb_trucks*pulp.lpSum([lhcost(da) for da in Useful_Da])

# Create Constraint : every Zip is allocated
print("Create contraint 'every zipcode is assigned to a DA'")
for pc in tqdm(Arcs.keys()):          
    prob += pulp.lpSum([Arcs[pc][da]['variable'] for da in Arcs[pc].keys()]) == 1

# Volume only if DC open, limit the max number of DA
for da in Useful_Da:
    Zip_temp = []
    for pc in Arcs.keys():
        try:
            Zip_temp.append(Arcs[pc][da]['variable'])
        except:
#            Whatever lline to prevent error in case da is not in zip dict
            j=10      
    prob += pulp.lpSum(Zip_temp)-1500*DAC_ZipCode_Dict[da]['opening_variable'] <= 0
    
 # Constraint over the weight variable   
for da in Useful_Da:
    prob += DAC_ZipCode_Dict[da]["Weight_variable"] >= 0
    prob += DAC_ZipCode_Dict[da]["Weight_variable"] >= pulp.lpSum([ZipCode_Dict[pc]['Volume']*Arcs[pc][da]['variable'] for pc in Arcs.keys() if da in Arcs[pc]]) / nb_trucks * weight_per_volume -weight_treshold_ltl 

# The problem is solved using PuLP's choice of Solver
print("Solve Problem")
start_time = time.clock()
solve= pulp.solvers.GUROBI(timeLimit = 300)

solve.actualSolve(prob)
end_time = time.clock()
print(end_time-start_time)

# The status of the solution is printed to the screen
print("Status:", pulp.LpStatus[prob.status])


# The optimised objective function value is printed to the screen    
print ("total cost", pulp.value(prob.objective))

#Create workbook for results
w_result = xl.Workbook()
wresultassign = w_result.create_sheet('Optimization Results Assignment')
wresultda =  w_result.create_sheet('Optimization Results DA')
# export results on excel

print("Exporting Results")

wresultassign.cell(row=1,column=1).value= "ZipCode"
wresultassign.cell(row=1,column=2).value= "Carrier"
wresultassign.cell(row=1,column=3).value= "DaZipCode"
wresultassign.cell(row=1,column=4).value= 'DA and Carrier'
wresultassign.cell(row=1,column=5).value= 'Volume'
wresultassign.cell(row=1,column=6).value= 'Unit Cost'
# Print Results on excel
r=2
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        if Arcs[pc][da]['variable'].varValue !=0:
            wresultassign.cell(row=r,column=1).value= pc
            wresultassign.cell(row=r,column=2).value= da[6:]
            wresultassign.cell(row=r,column=3).value= da[:5]
            wresultassign.cell(row=r,column=4).value= da
            wresultassign.cell(row=r,column=5).value= ZipCode_Dict[pc]['Volume']            
            wresultassign.cell(row=r,column=6).value= Arcs[pc][da]['lm_cost']
            r+=1
            
            
wresultda.cell(row=1,column=1).value= "Da"
wresultda.cell(row=1,column=2).value= "Carrier"
wresultda.cell(row=1,column=3).value= "Da zip"
wresultda.cell(row=1,column=4).value= 'Volume'
wresultda.cell(row=1,column=5).value= 'lh cost'

# Print Results on excel
r=2

for da in Useful_Da:
    if DAC_ZipCode_Dict[da]['opening_variable'].varValue !=0:
        wresultda.cell(row=r,column=1).value= da
        wresultda.cell(row=r,column=2).value= da[6:]
        wresultda.cell(row=r,column=3).value= da[:5]
        wresultda.cell(row=r,column=5).value= (DAC_ZipCode_Dict[da]["Weight_variable"].varValue * Da_Dfc[da[:5]]['Global']['slope']+ Da_Dfc[da[:5]]['Global']['cost_opening'])*nb_trucks

        r+=1




print("Save File")

w_result.save("C:\HomeDepot_Excel_Files\Optimized.xlsx")

# Return List of useful DA
Useful_Da = []
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        if Arcs[pc][da]['variable'].varValue != 0:
            Useful_Da = list(set().union([da],Useful_Da))
print('Number of useful DA :', len(Useful_Da))
         