# -*- coding: utf-8 -*-
"""
Created on Fri Aug  4 14:42:10 2017

@author: Bastide
"""

def geocode(postal, *List, recursion=0):
#    Returns the list[City,Postal Code,(lat,long)
    from geopy.exc import GeocoderTimedOut
    from geopy.geocoders import Nominatim
    import uszipcode as usz
    import time
#    Try with uszipcode
    search = usz.ZipcodeSearchEngine()
    info= search.by_zipcode(str(postal))
    
    if info["City"] is not None:
        city = info["City"]
        lat = info["Latitude"]
        long = info["Longitude"]  
        return [city,
                postal,
                (lat,long)]

#       if no result is found use geopy
    else:
        try:
            info = Nominatim().geocode(str(postal)+", United States of America")

            
            #       Avoid time out error
        except GeocoderTimedOut as e:
            if recursion > 10:      # max recursions
                raise e
            time.sleep(0.2) # wait a bit
            # try again
            return geocode(postal, List, recursion=recursion + 1)
            
            
        else: 
            # If no result found use previous info
            if info is None:
                print("attention")
                print(info)
                print(List[-1])
                if List is not None:
                    return List[-1]
                else:
                    return [None, None, (0,0)]

            else :
                city = info.address
                lat=info.latitude
                long=info.longitude
                return [city,
                    postal,
                    (lat,long)]

# Same without list for correction
def geocode2(postal, recursion=0):
#    Returns the list[City,Postal Code,(lat,long)
    from geopy.exc import GeocoderTimedOut
    from geopy.geocoders import GoogleV3
    import uszipcode as usz
    import time
#    Try with uszipcode
    search = usz.ZipcodeSearchEngine()
    info= search.by_zipcode(str(postal))
    
    if info["City"] is not None:
        city = info["City"]
        lat = info["Latitude"]
        long = info["Longitude"]  
        return [city,
                postal,
                (lat,long)]

#       if no result is found use geopy
    else:
        try:
            info = GoogleV3().geocode(str(postal)+", United States of America")

            
            #       Avoid time out error
        except GeocoderTimedOut as e:
            if recursion > 10:      # max recursions
                raise e
            time.sleep(0.1) # wait a bit
            # try again
            return geocode2(postal,recursion=recursion + 1)
            
            
        else: 
            # If no result found use previous info
            if info is None:
                return None

            else :
                city = info.address
                lat=info.latitude
                long=info.longitude
                return [city,
                    postal,
                    (lat,long)]


# Facilitate the openpyxl formating
def cell(Sheet, rownb, columnnb):  
    return Sheet.cell(row=rownb, column=columnnb).value

## Return distances between two postal codes
#def get_distance(Zip1,Zip2):
#    from geopy.distance import vincenty 
#    zip1=geocode2(Zip1)[2]
#    zip2=geocode2(Zip2)[2]
#    return vincenty(zip1,zip2).miles
#    


# Return the number of instance in the Sheet, we can adjust the starting line
def instance(Sheet, starting_row=2, column=1):
    r=0
    while cell(Sheet, r+starting_row, column) is not None:
        r+=1
    return r

    
# Return a dictionnary of pricing (State : Carrier : ( Flat , Break, Extra))
def get_lm_pricing(Sheet):
    Pricing={}
# Get State in DIct
    nb_state = instance(Sheet,starting_row=3)
    for r in range(nb_state):
        Pricing[cell(Sheet,3+r,2)]= {}
# Get carriers
    c=0
    Carriers = []
    while cell(Sheet,1,4+3*c) is not None:
        Carriers.append(cell(Sheet,1 ,4+3*c))
        c+=1
# Create Dictionnaries
    for r in range(len(Pricing)):
        for c in range(len(Carriers)):
#            Append only if there is pricing info
            if cell(Sheet,r+3,3*c+3) is not None:
                Pricing[cell(Sheet, r+3,2)][Carriers[c]]={'Flat':cell(Sheet,r+3,3+3*c),'Break':cell(Sheet,r+3,4+3*c),'Extra':cell(Sheet,r+3,5+3*c)}
    return Pricing

#    Return a list of all"neighboring" states to the one in the argument 
def neig_states(state_code, Sheet):
    a=instance(Sheet)
    List=[state_code]
    for r in range(a):
        if cell(Sheet,r+2,1) == state_code:
            List.append(cell(Sheet,r+2, 2))
            
        if cell(Sheet, r+2, 2) == state_code and cell(Sheet,r+2,3) == "1st":
            List.append(cell(Sheet, r+2, 1))
    return List
    
# Return a list of STates defining The region, based on degree of neighboor 
def get_second_neig(state_code, Sheet):
    List = neig_states(state_code, Sheet) 
    List2 = List
    for state in List:
        A = neig_states(state,Sheet)
#       Remove Duplicates
        List2 = list(set().union(List2,A))
    return List2          


# Compute Distance using zip database that get updated if new zip happens
def compute_distance(Workbook, Sheet, column_origin, column_destination, column_distance):
    import openpyxl as xl
    from geopy.distance import vincenty
    from Procedures import instance, cell, geocode2
    print('Open File')
    wb = xl.load_workbook(Workbook)
    wsdist = wb[Sheet]
    print('Open Database')
    wdata = xl.load_workbook('C:\HomeDepot_Excel_Files\Zip_latlong.xlsx')
    wslatlong = wdata['Zip']
    
    linelatlong = instance(wslatlong)
#    Collect Data
    Zip_lat_long = {}
    for r in range(linelatlong):
        Zip_lat_long[str(cell(wslatlong,r+2,1))] = (cell(wslatlong,r+2,2),cell(wslatlong,r+2,3))
        print ('Collecting Data ',r*100/linelatlong,'%')
     
#        Compute distance
    linedistance = instance(wsdist)
#   a serve to know if zipcode not in database appears
    a = 0
    for r in range(linedistance):
        try :
            wsdist.cell(row=r+2,column=column_distance).value = vincenty(Zip_lat_long[str(cell(wsdist,r+2,column_origin))],Zip_lat_long[str(cell(wsdist,r+2,column_destination))]).miles

        except KeyError:
            Zip_lat_long[str(cell(wsdist,r+2,column_origin))]= (geocode2(cell(wsdist,r+2,column_origin))[2][0],geocode2(cell(wsdist,r+2,column_origin))[2][1])
            Zip_lat_long[str(cell(wsdist,r+2,column_destination))]= (geocode2(cell(wsdist,r+2,column_destination))[2][0],geocode2(cell(wsdist,r+2,column_destination))[2][1])
            wsdist.cell(row=r+2,column=column_distance).value = vincenty(Zip_lat_long[str(cell(wsdist,r+2,column_origin))],Zip_lat_long[str(cell(wsdist,r+2,column_destination))]).miles
            a+=1
        print('Compute Distances ',(r+1)/linedistance*100, '%')
    print('Saving File')
    wb.save(Workbook)
#    Update database
    if a!=0:
        print("Update Database")
        ZipList = Zip_lat_long.keys()
        c = 0
        for r in ZipList:
            wslatlong.cell(row = c+2,column = 1).value = r
            wslatlong.cell(row = c+2,column = 2).value = Zip_lat_long[r][0]
            wslatlong.cell(row = c+2,column = 3).value = Zip_lat_long[r][1]
            c+=1
            print('Updating Database ',c*100/len(ZipList),'%')
        wdata.save('C:\HomeDepot_Excel_Files\Zip_latlong.xlsx')
        print('Database updated')
        
def compute_distance2(zip1,zip2,Dict_lat_long):
    b = 0
    from geopy.distance import vincenty
    from Procedures import geocode2    
#   a serve to know if zipcode not in database appears and will
    try:
        latlong1 = Dict_lat_long[zip1]
    except KeyError:
        Dict_lat_long[zip1] = geocode2(zip1)[2]
        latlong1 = Dict_lat_long[zip1]
        b=1
        
    try:
        latlong2 = Dict_lat_long[zip2]
    except KeyError:
        Dict_lat_long[zip2] = geocode2(zip2)[2]
        latlong2 = Dict_lat_long[zip2]    
        b=1
        
    distance = vincenty(latlong1,latlong2).miles
    
    return distance, Dict_lat_long ,b

          
# This function will return a 5 digit postal code by adding 0 in front if the input is less than 5
def correct_zip(str_Zip):
    if len(str_Zip) == 4:
        zipcode = "0"+str_Zip
    elif len(str_Zip) == 3:
        zipcode = "00"+str_Zip
    elif len(str_Zip)>5:
        a = len(str_Zip)-5
        zipcode = str_Zip[a:]
    else :
        zipcode = str_Zip
    return zipcode
    