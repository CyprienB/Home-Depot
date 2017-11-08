# -*- coding: utf-8 -*-
"""
Created on Fri Sep 29 14:53:36 2017
This file combine all other Python files with some improvments to avoid useless processes
This is an optimization of the Home Depot .com Delivery network, with objective function being the total cost (line haul + last Mile) 
@author: Cyprien Bastide, Steven (Gao) Ming, Edson David Silva Moreno
"""

# This module is used to import the different modules that will be used in this file
import sys
# This module is used to open excel files
import openpyxl as xl
# This module is used to perform regression analysis
from statsmodels.formula.api import ols
# This module is used to create progress bar.
from tqdm import tqdm
# This module is the optimization engine
import pulp
# This module is used to compute time elapsed for a task
import time
# This module is used to convert spreadsheet into a specific dataframe for analysis
import pandas as pd
# neig_states: it returns the neighboring state of the input state
# compute_distance2: it computes the distances from 2 zip code and the lat long database
# correct_zip: it adds 0 in front of postal codes that are not 5 digits long
# get_lm_pricing: it returns a dictionnary containing the info to compute the last mile cost
# averageOrig: # it returns the dictionary of every State Destination with weighted origin 
from Procedures import neig_states, compute_distance2, correct_zip, get_lm_pricing, averageOrig

# Define and initialize fixed variables
number_days = 30*6
weight_treshold_ltl = 200
nb_trucks = round(number_days*5/7)
weight_per_volume = 100

# Import and convert spreadsheets into panda dataframe
wb = pd.ExcelFile('C:\HomeDepot_Excel_Files\Standard_File.xlsx')
w_neig = wb.parse('List_of_Neighboring_States')
w_da = wb.parse('DA_List')
w_zip = wb.parse('Zip_Allocation_and_Pricing')
w_range = wb.parse("Zip_Range")
w_dfc = wb.parse("DFC list")
ltl_price = wb.parse('ltl_price', converters={'dest_zip': str,'orig_zip': str})

# Import Database of Zipcode Latitude and Longitude
print('Open Lat Long Database')
wdata = pd.ExcelFile('C:\HomeDepot_Excel_Files\Zip_latlong.xlsx')
wslatlong = wdata.parse('Zip')

# Parse useful data for regression analysis
ltl_price = ltl_price[(ltl_price['tot_shp_wt'] >= 200) & (ltl_price['tot_shp_wt'] <= 4999) & (ltl_price['aprv_amt'] <=1000)]
# Define an interaction term between distance and weight
ltl_price['tot_mile_wt'] = ltl_price['tot_mile_cnt'] * ltl_price['tot_shp_wt']

# Fit this regression model with .fit() and show results
linehaul_model = ols('aprv_amt ~ tot_mile_cnt + tot_shp_wt + + tot_mile_wt', data=ltl_price).fit()

# Output the result of the regression model
linehaul_model_summary = linehaul_model.summary()
print(linehaul_model_summary)

# Store significant terms from the regression model
variables = [linehaul_model.params.index.tolist()][0]

# Filter and rename orig_state if 'orig_state' variable is used in the regression model
for i in range(0, len(variables)):
    if variables[i].find("orig_state") != -1:
        variables[i] = variables[i][-3]+variables[i][-2]

# Store coefficents to the significant terms from the regression model
coeff = [linehaul_model.params.tolist()][0]

# Convert two lists (significant terms & coefficents) to a dictionary
coefficient = dict(zip(variables,coeff))

"""
###############################################################
###############################################################

This part create the different dictionnaries that are going to be used to solve the optimization problem
Multiple dictionnaries are going to be created to represent DAs and Zipcode because we will need different keys to call them

###############################################################
###############################################################
"""
print ("Create all Dictionnaries")
# Get length of DA, Zip, latlong, fullfillment centers, zip range
n_da= len(w_da)
n_zip = len(w_zip)
linelatlong = len(wslatlong)
nbdfc = len(w_dfc)
n_range = len(w_range)

# Dictionnary for Zip {State : [( zipcode, volume)]}
State_Zip_dict = {}
#Create Dictionnary for Zipcode (Zip, Volume ,State)
ZipCode_Dict={}
for r in range(n_zip):
    state = w_zip['Zip State'][r]
    zipcode = correct_zip(str(w_zip['Zip#'][r])) 
    volume = w_zip['Volume'][r]
    State_Zip_dict.setdefault(state, []).append((zipcode, volume))
    ZipCode_Dict[zipcode]={'Zip':zipcode,'Volume':volume, 'State':state}

# Dictionnary for DA  {State : [Da_ZipCode]}. This is useful because Arcs are going to be created based on neighgboring states
State_Da_dict = {}   
# Dictionnary for Da { Zip:(Zip, State, [Carrier])} will be useful to compute distances       
DA_ZipCode_Dict = {}     
# Other dictionnary for Da, { Zip + Carrier : (Zip, State, Carrier)}   
DAC_ZipCode_Dict = {}       

for r in range(n_da):
    zipcode = correct_zip(str(w_da['Zip_Code'][r]))
    state = w_da['State'][r]
    carrier = w_da['Carrier'][r]
    # Remove duplicate: DA in same zipcode but different carriers
    State_Da_dict[state] = list(set().union(State_Da_dict.setdefault(state,[]),[zipcode]))
    DA_ZipCode_Dict.setdefault(zipcode,{'Zip':zipcode, 'State':state, 'Carrier':[carrier]}) 
    DA_ZipCode_Dict[zipcode]['Carrier'] = list(set().union(DA_ZipCode_Dict[zipcode]['Carrier'],[carrier]))
    DAC_ZipCode_Dict[zipcode+' '+ carrier] = {'Zip':zipcode, 'State':state, 'Carrier':carrier}

# Create dictionnary for the database lat long{Zip : (lat,long)}
Zip_lat_long = {}
for r in range(linelatlong):
    zipcode = correct_zip(str(wslatlong['Zip#'][r]))
    lat = wslatlong['lat'][r]
    long = wslatlong['long'][r]
    Zip_lat_long[zipcode] = (lat,long)

# Create Dictionnary of DFC {State : {Name, Zip, State}}
DFC_Dict={}
for r in range(nbdfc):
    name = w_dfc['DFC'][r]
    zipcode = correct_zip(str(w_dfc['DFC ZIP'][r]))
    state = w_dfc['DFC State'][r]
    DFC_Dict[state]={'State':state,'Name':name,'Zipcode':zipcode}

# Import LM Pricing again using openxl
wx = xl.load_workbook('C:\HomeDepot_Excel_Files\Standard_File.xlsx')
w_lm_pricing = wx["LM_Pricing"]

# Create LM Pricing Dictionnary
Pricing = get_lm_pricing(w_lm_pricing)

# Get arc max range Dictionnary {State : Max range}
Range = { w_range['Abreviation'][r] : w_range['Maximum distance between Zip (in state) to DA (out of state)'][r] for r in range(n_range)}

# Create dictionary with State Destination as a Key and nested dictionary with weight by origin 
percDestin = averageOrig(ltl_price)
    
"""
###############################################################
###############################################################

This part compute the distance between Das and DFC and the pricing based on parameters and distance

\\ Problem for new DAs in States without Das currently (Key Error)

###############################################################
############################################################### 
"""
# Da_Dfc format {da : {state_dfc : {distance, percentage from state}, global : {slope,cost_opening}}
a = 0
Da_Dfc = {}

for da in DA_ZipCode_Dict.keys():
    da_state = DA_ZipCode_Dict[da]["State"]
    try:
        for dfc_state in percDestin[da_state].keys():
            dfc_zip = DFC_Dict[dfc_state]['Zipcode']
            percentage = percDestin[da_state][dfc_state]
            distance, Zip_lat_long, b = compute_distance2(da,dfc_zip,Zip_lat_long)
            slope = coefficient["tot_mile_wt"]+coefficient["tot_mile_wt"]*distance
            intercept = coefficient['Intercept']+coefficient['tot_mile_cnt']*distance
            cost_opening = intercept + weight_treshold_ltl * slope
            Da_Dfc.setdefault(da,{}).setdefault(dfc_state, {"distance":distance, "percentage" : percentage,"slope": slope,"cost_opening": cost_opening})
            if b == 1 :
                a += 1
        slope = sum(Da_Dfc[da][dfc_state]["slope"]*Da_Dfc[da][dfc_state]["percentage"] for dfc_state in Da_Dfc[da].keys())
        cost_opening = sum(Da_Dfc[da][dfc_state]["cost_opening"]*Da_Dfc[da][dfc_state]["percentage"] for dfc_state in Da_Dfc[da].keys())
        Da_Dfc[da]['Global']={'slope' : slope, 'cost_opening' : cost_opening}
    except KeyError:
#        Just assume we take the previous cost
        Da_Dfc.setdefault(da,{}).setdefault('Global',{'slope' : slope, 'cost_opening' : cost_opening, 'Warning':"This Da doesn't have real slope or cost of opening"}  )  
"""
###############################################################
###############################################################

This part create the combination DA Zipcode based on neighboring states and then compute the distances 

###############################################################
###############################################################
"""

# Create couple that need to have distance assigned to it based Neighboring states, 
# Go through every state, look at the Da inside, and assign them to all zip code in neighboring states
combination = []
# Iterate through the states
print('Creating combination Da Zipcode')
for state in tqdm(State_Da_dict.keys()):
#    Create list of Da in the state
    da_list = State_Da_dict[state]

#   Create list of Zipcode that are in neigh state
    zip_list = []
    neighboring_states= neig_states(state,w_neig)
    for n_state in neighboring_states:
        zip_list += State_Zip_dict[n_state]
# Create combination if Volume is not 0, have list of all combination [[da,zip,distance]]
    for da in da_list:
        for pc in zip_list:
            zipcode = pc[0]
            volume = pc[1]
            distance, Zip_lat_long, b = compute_distance2(da, zipcode, Zip_lat_long)
            combination.append([da,zipcode,distance])
            if b == 1 :
                a += 1

# Update if new zipcodes have been added
if a != 0:
    print("Update Database")
    ZipList = Zip_lat_long.keys()
    c = 0
    for r in ZipList:
        wslatlong.cell(row = c+2,column = 1).value = r
        wslatlong.cell(row = c+2,column = 2).value = Zip_lat_long[r][0]
        wslatlong.cell(row = c+2,column = 3).value = Zip_lat_long[r][1]
        c+=1
    wdata.save('C:\HomeDepot_Excel_Files\Zip_latlong.xlsx')
    print('Database updated')
  
"""
###############################################################
###############################################################

This part is the HDU optimization model (includes last mile and line haul)
and uses all DA(based on previous optimization)

\For now arcs are recreated while we could use previous dictionnary
\Model only has one treshold
###############################################################
###############################################################
"""
# Create dictionnary for the arcs
Arcs={}
for i in tqdm(range(len(combination))):
    da = combination[i][0]
    zipcode = combination[i][1] 
    distance = combination[i][2]
#        Create an arc only if distance between DA and Zip is less than the Zip's state threshold in this model we only use volume above zero
    if distance< Range[ZipCode_Dict[zipcode]['State']] and ZipCode_Dict[zipcode]['Volume']>0:
        for carrier in DA_ZipCode_Dict[da]['Carrier']:
            Arcs.setdefault(zipcode,{}).setdefault(da +" "+carrier,{'distance' : distance})
                
                
# Compute Costs for the arcs
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        
        distance = Arcs[pc][da]['distance']        
        da_state = DAC_ZipCode_Dict[da]['State']
        da_carrier = DAC_ZipCode_Dict[da]['Carrier']
        
        try:
            flat = Pricing[da_state][da_carrier]['Flat']
            breakpoint = Pricing[da_state][da_carrier]['Break']
            extra = Pricing[da_state][da_carrier]['Extra']
        except KeyError:
            sys.exit(("LM_Cost spreadsheet does not contain pricing info for couple state-carrier %s, %s, need to update Standard File" %(da_state,da_carrier)))
            
#        Check if distance Da_Zip is within flat distance
        if distance <  breakpoint:
            lmcost=flat
        else :
            lmcost=flat+ (distance - breakpoint) * extra
        if distance > 75 : # Opportunity cost
            lmcost += 25
        Arcs[pc][da]['lm_cost'] = lmcost

#   { Zip : {DA+Carrier :{distance, lm_cost (come in next step), var(come in two steps)}}
print("Create Model")

# Create Model
prob = pulp.LpProblem("Minimize HDU Cost",pulp.LpMinimize)

# Design arcs
for pc in Arcs.keys():
    for da in Arcs[pc].keys():

        var = pulp.LpVariable("Arc_%s_%s)" % (pc,da),0,1,pulp.LpContinuous)
        Arcs[pc][da]['variable']=var
            
# Create variable for Das (OPEN AND VOLUME OVER 200)
for da in DAC_ZipCode_Dict.keys():

    ovar = pulp.LpVariable("Da_%s" % (str(da)),0,1,pulp.LpBinary)
    wvar = pulp.LpVariable("Da_%s_above_200LBS" % (str(da)))
    DAC_ZipCode_Dict[da]['opening_variable']=ovar
    DAC_ZipCode_Dict[da]['Weight_variable']=wvar


# Create Objective function : minimize cost
print("Create objective and Constraint")
def lmcost(pc,da):
    return Arcs[pc][da]['lm_cost']*ZipCode_Dict[pc]['Volume']*Arcs[pc][da]['variable']
def lhcost(da):
    zip_da = da[:5]
    return Da_Dfc[zip_da]['Global']['cost_opening']*DAC_ZipCode_Dict[da]['opening_variable'] + Da_Dfc[zip_da]['Global']['slope'] * DAC_ZipCode_Dict[da]["Weight_variable"]

prob += pulp.lpSum([lmcost(pc,da) for pc in Arcs.keys() for da in Arcs[pc].keys()]) + nb_trucks*pulp.lpSum([lhcost(da) for da in DAC_ZipCode_Dict.keys()])

# Create Constraint : every Zip is allocated
print("Create contraint 'every zipcode is assigned to a DA'")
for pc in tqdm(Arcs.keys()):          
    prob += pulp.lpSum([Arcs[pc][da]['variable'] for da in Arcs[pc].keys()]) == 1

# Volume only if DA open, limit the max number of Zip
for da in DAC_ZipCode_Dict.keys():
    Zip_temp = []
    for pc in Arcs.keys():
        try:
            Zip_temp.append(Arcs[pc][da]['variable'])
        except:
#            Whatever lline to prevent error in case da is not in zip dict
            j=10      
    prob += pulp.lpSum(Zip_temp)-1500*DAC_ZipCode_Dict[da]['opening_variable'] <= 0
    
 # Constraint over the weight variable   
for da in DAC_ZipCode_Dict.keys():
    prob += DAC_ZipCode_Dict[da]["Weight_variable"] >= 0
    prob += DAC_ZipCode_Dict[da]["Weight_variable"] >= pulp.lpSum([ZipCode_Dict[pc]['Volume']*Arcs[pc][da]['variable'] for pc in Arcs.keys() if da in Arcs[pc]]) / nb_trucks * weight_per_volume -weight_treshold_ltl 

# Open a certain number of DA
#    
#prob += pulp.lpSum([DAC_ZipCode_Dict[da]['opening_variable'] for da in DAC_ZipCode_Dict.keys() ])>= 120

# The problem is solved using PuLP's choice of Solver
print("Solve Problem")
start_time = time.clock()
solve= pulp.solvers.GUROBI(timeLimit = 500)

solve.actualSolve(prob)
end_time = time.clock()
print(end_time-start_time)

# The status of the solution is printed to the screen
print("Status:", pulp.LpStatus[prob.status])


# The optimised objective function value is printed to the screen    
print ("total cost", pulp.value(prob.objective))

"""
#########################
Put results into dataframe
##########################
"""

print("Exporting Results")

column_names=["ZipCode",
         "Carrier",
         "DaZipCode",
         'DA and Carrier',
         'Volume',
         'Unit Cost',
         'Total Cost',
         'Assignment Variable',
         "distance"]
# Print Results as DataFrame
Assign_Results = []
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        if Arcs[pc][da]['variable'].varValue > 0.001 :
            Assign_Results.append([pc,
                                   da[6:],
                                   da[:5],
                                   da,
                                   ZipCode_Dict[pc]['Volume'],
                                   Arcs[pc][da]['lm_cost'],
                                   Arcs[pc][da]['lm_cost'] * Arcs[pc][da]['variable'].varValue*ZipCode_Dict[pc]['Volume'],
                                   Arcs[pc][da]['variable'].varValue,
                                   Arcs[pc][da]['distance']])
Assign_Results = pd.DataFrame(Assign_Results,columns = column_names)
            
            
column_names_da= ["Da",
              "Carrier",
              "Da zip",
              'Volume above 200',
              'lh cost',
              'Opening_variable']

DA_Results = []
Useful_Da = []
for da in DAC_ZipCode_Dict.keys():
    if DAC_ZipCode_Dict[da]['opening_variable'].varValue > 0.0001:
        DA_Results.append([da,
                           da[6:],
                           da[:5],
                           DAC_ZipCode_Dict[da]["Weight_variable"].varValue,
                           (DAC_ZipCode_Dict[da]["Weight_variable"].varValue * Da_Dfc[da[:5]]['Global']['slope']+ Da_Dfc[da[:5]]['Global']['cost_opening'])*nb_trucks,
                            DAC_ZipCode_Dict[da]["opening_variable"].varValue])
 # Return List of useful DA       
        Useful_Da = list(set().union([da],Useful_Da))
        
        r+=1
DA_Results = pd.DataFrame(DA_Results,columns = column_names_da)


print("Number of Useful Das:", len(Useful_Da))


#print("Save File")
#
#w_result.save("C:\HomeDepot_Excel_Files\Optimized.xlsx")
         


"""
#####################################
#####################################
Optimize 0 Volume Postal and leftovers
#####################################
#####################################
"""
#
# Create dictionnary for the arcs
Arcs0={}
for i in tqdm(range(len(combination))):
    da = combination[i][0]
    zipcode = combination[i][1] 
    distance = combination[i][2]

    if zipcode not in Arcs.keys():
        for carrier in DA_ZipCode_Dict[da]['Carrier']:
            if da + " " + carrier in Useful_Da:
                Arcs0.setdefault(zipcode,{}).setdefault(da +" "+carrier,{'distance' : distance})
                
                
# Compute Costs for the arcs
for pc in Arcs0.keys():
    for da in Arcs0[pc].keys():
        
        distance = Arcs0[pc][da]['distance']        
        da_state = DAC_ZipCode_Dict[da]['State']
        da_carrier = DAC_ZipCode_Dict[da]['Carrier']
        
        try:
            flat = Pricing[da_state][da_carrier]['Flat']
            breakpoint = Pricing[da_state][da_carrier]['Break']
            extra = Pricing[da_state][da_carrier]['Extra']
        except KeyError:
            sys.exit(("LM_Cost spreadsheet does not contain pricing info for couple state-carrier %s, %s, need to update Standard File" %(da_state,da_carrier)))
            
#        Check if distance Da_Zip is within flat distance
        if distance <  breakpoint:
            lmcost=flat
        else :
            lmcost=flat+ (distance - breakpoint) * extra
        if distance > 75 : # Opportunity cost
            lmcost += 25
        Arcs0[pc][da]['lm_cost'] = lmcost

#   { Zip : {DA+Carrier :{distance, lm_cost (come in next step), var(come in two steps)}}
print("Create Model")

# Create Model
prob = pulp.LpProblem("Minimize LM Cost",pulp.LpMinimize)

# Design arcs
for pc in Arcs0.keys():
    for da in Arcs0[pc].keys():

        var = pulp.LpVariable("Arc_%s_%s)" % (pc,da),0,1,pulp.LpContinuous)
        Arcs0[pc][da]['variable']=var
            

# Create Objective function : minimize cost
print("Create objective and Constraint")
def lmcost2(pc,da):
    return (Arcs0[pc][da]['lm_cost']+0.01*Arcs0[pc][da]['distance'])*Arcs0[pc][da]['variable']

prob += pulp.lpSum([lmcost2(pc,da) for pc in Arcs0.keys() for da in Arcs0[pc].keys()])

# Create Constraint : every Zip is allocated
print("Create contraint 'every zipcode is assigned to a DA'")
for pc in tqdm(Arcs0.keys()):          
    prob += pulp.lpSum([Arcs0[pc][da]['variable'] for da in Arcs0[pc].keys()]) == 1

# The problem is solved using PuLP's choice of Solver
print("Solve Problem")

solve= pulp.solvers.GUROBI(timeLimit = 300)

solve.actualSolve(prob)

"""
#########################
Put results into dataframe
##########################
"""

print("Exporting Results")

column_names=["ZipCode",
         "Carrier",
         "DaZipCode",
         'DA and Carrier',
         'Volume',
         'Unit Cost',
         'Total Cost',
         'Assignment Variable',
         "distance"]
# Print Results as DataFrame
Assign_Results2 = []
for pc in Arcs0.keys():
    for da in Arcs0[pc].keys():
        if Arcs0[pc][da]['variable'].varValue > 0.001 :
            Assign_Results2.append([pc,
                                   da[6:],
                                   da[:5],
                                   da,
                                   ZipCode_Dict[pc]['Volume'],
                                   Arcs0[pc][da]['lm_cost'],
                                   Arcs0[pc][da]['lm_cost'] * Arcs0[pc][da]['variable'].varValue*ZipCode_Dict[pc]['Volume'],
                                   Arcs0[pc][da]['variable'].varValue,
                                   Arcs0[pc][da]['distance']])
Assign_Results = Assign_Results.append(pd.DataFrame(Assign_Results2,columns = column_names), ignore_index=True)
            
print("Write Excel")
writer = pd.ExcelWriter('C:\HomeDepot_Excel_Files\Optimized_oportunity.xlsx', engine='xlsxwriter')
Assign_Results.to_excel(writer,'AssignmentResults')
DA_Results.to_excel(writer,'OptimizedDA')
writer.save()



