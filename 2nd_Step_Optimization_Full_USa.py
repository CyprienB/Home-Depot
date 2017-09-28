# -*- coding: utf-8 -*-
"""
Created on Sun Aug  6 09:39:37 2017
Test for Github
@author: Bastide
"""

# This is the optimization file
import time
import openpyxl as xl
from Procedures import cell, instance, get_lm_pricing
import pulp
# Open Excel worksheet

print("Open Worksheet")
start_time = time.clock()
wb = xl.load_workbook("C:\HomeDepot_Excel_Files\File_modified.xlsx")

# Sheet with DA info
wda = wb["DA_List"]

# Sheet with Zip info
wzip = wb["Zip_Allocation_and_Pricing"]

# Sheet with Distance
wdistance = wb["Distances"]
end_time = time.clock()
print(end_time-start_time)


#Get info from DA { Zip:(Zip, City, State, [Carrier])}
start_time = time.clock()
print("Importing info")
numberDa = instance(wda)
Da = {}
for r in range(numberDa):
    try :
        Da[str(cell(wda, 2+r,2))]['Carrier'].append(cell(wda,r+2,4))
    except KeyError:
        Da[str(cell(wda, 2+r,2))] = {'Zip':str(cell(wda, 2+r,2)),'City':cell(wda, 2+r, 1), 'State':cell(wda, 2+r, 3), 'Carrier':[cell(wda, 2+r, 4)]}


#Get info from Zipcode (Zip, Volume , City, State)
numberZip = instance(wzip)
Zip={}
for r in range(numberZip):
    zi = str(cell(wzip, 2+r, 1))
    if len(zi)==4:
        zi = '0'+zi
    Zip[zi]={'Zip':zi,'Volume':cell(wzip, 2+r,7), 'City':cell(wzip, 2+r, 3), 'State':cell(wzip, 2+r, 2)} 



# Get pricing info
Pricing=get_lm_pricing('C:\HomeDepot_Excel_Files\Last_Mile_Pricing.xlsx','Sheet1')

# Get arc max range {State : Max range}
wrange = wb['Zip_Range']
Range = { cell(wrange,r+2,2) : cell(wrange,r+2,3) for r in range(instance(wrange))}
    

#Get Arcs and Distance { Zip : {DA+Carrier :{distance, lm_cost (come in next step), var(come in two steps)}}
print("Build Arcs")
start_time = time.clock()
Arcs={}
linedistance = instance(wdistance)
for r in range(linedistance):
    dat = str(cell(wdistance,r+2,1))
    zipcodet = str(cell(wdistance, r+2,2))
    distance = cell(wdistance,r+2,3)
#    Deal with 4 digit Zip
    if len(dat) == 4:
        dat = "0"+dat
    if len(zipcodet)==4:
        zipcodet = "0"+zipcodet
#        Create an arc only if distance between DA and Zip is less than the Zip's state threshold
    if distance< Range[Zip[zipcodet]['State']]:
        try:
            for carrier in Da[dat]['Carrier']:
                Arcs[zipcodet][dat +" "+carrier ]={'distance' : cell(wdistance,r+2,3)}
        except KeyError:
            Arcs[zipcodet]= {}
            for carrier in Da[dat]['Carrier']:
                 Arcs[zipcodet][dat + " "+carrier]={'distance' : cell(wdistance,r+2,3)}

# Compute Costs for the arcs
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        distance = Arcs[pc][da]['distance']        
        da_state = Da[da[:5]]['State']
        da_carrier = da[6:]
        flat = Pricing[da_state][da_carrier]['Flat']
        breakpoint = Pricing[da_state][da_carrier]['Break']
        extra = Pricing[da_state][da_carrier]['Extra']
#        Check if distance Da_Zip is within flat distance
        if distance <  breakpoint:
            Arcs[pc][da]['lm_cost']=flat
#       Compute cost if above breakpoint
        else :
            Arcs[pc][da]['lm_cost']=flat+ (distance - breakpoint) * extra
        
                
# Create Model
prob = pulp.LpProblem("Minimize Distance",pulp.LpMinimize)

# Design arcs
arc_assign=[]
cost_arc=[]
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        var = pulp.LpVariable("Arc_%s_%s)" % (str(pc),str(da)),0,1,pulp.LpContinuous)
        Arcs[pc][da]['variable']=var

end_time = time.clock()
print(end_time-start_time) 
 

# Create Objective function : minimize distance
print("Create objective and Constraint")
start_time = time.clock()
#           We add a fraction of distance to the lm cost so we can avoid equality in price
prob += pulp.lpSum([(Arcs[pc][da]['lm_cost']+0.001*Arcs[pc][da]['distance'])*Arcs[pc][da]['variable'] for pc in Arcs.keys() for da in Arcs[pc].keys()])

# Create Constraint : every Zip is allocated
for pc in Arcs.keys():          
    prob += pulp.lpSum([Arcs[pc][da]['variable'] for da in Arcs[pc].keys()]) == 1
end_time = time.clock()
print(end_time-start_time)

##   Avoid da with one zipcode
#for da in Da.keys():
#    for carrier in Da[da]['Carrier']:
#        prob += pulp.lpSum([Arcs[pc][da]['variable']] for pc in Arcs.keys()) != 1

# The problem is solved using PuLP's choice of Solver
print("Solve Problem")
start_time = time.clock()
prob.solve()
end_time = time.clock()
print(end_time-start_time)

# The status of the solution is printed to the screen
print("Status:", pulp.LpStatus[prob.status])


# The optimised objective function value is printed to the screen    
print ("choices ", pulp.value(prob.objective))

#Create or not the worksheet for results
try: 
    wresult = wb["Optimization_Results"]
except KeyError:
    wb.create_sheet("Optimization_Results")
    wresult = wb["Optimization_Results"]

# export results on excel

print("Exporting Results")
start_time = time.clock()
wresult.cell(row=1,column=1).value= "ZipCode"
wresult.cell(row=1,column=2).value= "Carrier"
wresult.cell(row=1,column=3).value= "DaZipCode"
wresult.cell(row=1,column=4).value= 'DA and Carrier'
wresult.cell(row=1,column=5).value= 'Volume'
wresult.cell(row=1,column=6).value= 'Unit Cost'
# Print Results on excel
r=2
for pc in Arcs.keys():
    for da in Arcs[pc].keys():
        if Arcs[pc][da]['variable'].varValue !=0:
            wresult.cell(row=r,column=1).value= pc
            wresult.cell(row=r,column=2).value= da[6:]
            wresult.cell(row=r,column=3).value= da[:5]
            wresult.cell(row=r,column=4).value= da
            wresult.cell(row=r,column=5).value= Zip[pc]['Volume']            
            wresult.cell(row=r,column=6).value= Arcs[pc][da]['lm_cost']
            r+=1
end_time = time.clock()
print(end_time-start_time)

print("Save File")
start_time = time.clock()
wb.save("C:\HomeDepot_Excel_Files\Optimized.xlsx")
end_time = time.clock()
print(end_time-start_time)

