# -*- coding: utf-8 -*-
"""
Created on Sat Aug  5 15:30:32 2017

@author: Bastide
"""

import openpyxl as xl
from Procedures import cell, instance, geocode2


wb = xl.load_workbook("Excel Files/Zip_latlong.xlsx")
ws = wb["Zip"]
line = instance(ws)
Latlong={cell(ws,r+2,1):{"lat":cell(ws,r+2,4),"long":cell(ws,r+2,5)} for r in range(line)}

wb = xl.load_workbook("File_Modified.xlsx")
ws = wb["Zip_lat_long"]

line = instance(ws)
for r in range(line):
    try:
        ws.cell(row=r+2,column=2).value= Latlong[str(cell(ws,r+2,1))]["lat"] 
        ws.cell(row=r+2,column=3).value= Latlong[str(cell(ws,r+2,1))]["long"] 
    except KeyError:
        postal = geocode2(cell(ws,r+2,1))
        ws.cell(row=r+2,column=2).value= postal[2][0]
        ws.cell(row=r+2,column=3).value= postal[2][1]
        print(cell(ws,r+2,1))

        
wb.save("File_Modified.xlsx")