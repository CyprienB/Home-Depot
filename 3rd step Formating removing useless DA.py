# -*- coding: utf-8 -*-
"""
Created on Wed Sep  6 20:59:49 2017

@author: Cyprien
"""


import openpyxl as xl
from Procedures import neig_states, cell, instance, compute_distance
# Open Worksheet
wb = xl.load_workbook('Excel Files\Optimized.xlsx')
# open Sheets and create Distances sheet
w_neig = wb['List_of_Neighboring_States']
w_da = wb['DA_List']
w_zip = wb['Zip_Allocation_and_Pricing']
w_dis= wb['Distances']
wb.remove_sheet(w_dis)
wb.create_sheet('Distances')
w_dis = wb['Distances']
w_opti = wb['Optimization_Results']

#Create list of useful DA zip code based on optimization 1 result
useful_DA_nb = instance(w_opti,2,3)
useful_DA = []
for r in range(useful_DA_nb):
    useful_DA = list(set().union([str(cell(w_opti,r+2,3))],useful_DA)) 

w_dis.cell(row=1,column=1).value = 'DA'
w_dis.cell(row=1,column=2).value = 'ZipCode'
w_dis.cell(row=1,column=3).value = 'Distance DA-Zip'

# Get number of DA and of Zip
n_da= instance(w_da)
n_zip = instance(w_zip)

# Create dictionnaries for DA and Zip(with volume in tuple)
DA_dict = {}
Zip_dict = {}

for r in range(n_da):
    if str(cell(w_da,r+2,2)) in useful_DA:
        try:
    #        Remove duplicates, DA in same pc but different carriers
            DA_dict[cell(w_da,r+2,3)] = list(set().union(DA_dict[cell(w_da,r+2,3)],[int(cell(w_da,r+2,2))]))
        except KeyError:
            DA_dict[cell(w_da,r+2,3)]=[]
            DA_dict[cell(w_da,r+2,3)].append(int(cell(w_da,r+2,2)))
        
for r in range(n_zip):
    try:
        Zip_dict[cell(w_zip,r+2,2)].append((int(cell(w_zip,r+2,1)), cell(w_zip, r+2,7)))
    except KeyError:
        Zip_dict[cell(w_zip,r+2,2)]=[]
        Zip_dict[cell(w_zip,r+2,2)].append((int(cell(w_zip,r+2,1)), cell(w_zip, r+2,7)))
        
# Create couple that need to have distance assigned to it
r=2
rr=2
save =[]
c = 0
for state in DA_dict.keys():
    compt=0

    print(state)
    da_list = DA_dict[state]
    zip_list = []
    region= neig_states(state,w_neig)
    for region_state in region:
        zip_list += Zip_dict[region_state]


    for da in da_list:
        for pc in zip_list:
            if pc[1] != 0:
                if r<=1048576:
                    w_dis.cell(row=r,column=1).value = str(da)
                    w_dis.cell(row=r,column=2).value = pc[0]  
                    r +=1 
                else :
                    w_dis.cell(row=rr,column=5).value = str(da)
                    w_dis.cell(row=rr,column=6).value = pc[0]
                    rr+=1

# Save file
print ("save file")
wb.save('Excel Files\File_modified_2.xlsx')

#Compute distances
print("Compute Distances")
compute_distance('Excel Files\File_modified_2.xlsx','Distances',1,2,3)
